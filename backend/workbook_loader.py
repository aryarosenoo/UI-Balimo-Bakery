from __future__ import annotations

import re
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "Data TA"

PRODUCT_COLORS = [
    "#f97316",
    "#fb7185",
    "#facc15",
    "#2dd4bf",
    "#818cf8",
    "#38bdf8",
    "#22c55e",
    "#e879f9",
    "#94a3b8",
]

DAY_COLORS = {
    "Senin": "#3b82f6",
    "Selasa": "#8b5cf6",
    "Rabu": "#22c55e",
    "Kamis": "#f59e0b",
    "Jumat": "#ef4444",
    "Sabtu": "#ec4899",
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", normalize_text(value).lower())


def slugify(value: Any) -> str:
    text = normalize_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text or "item"


def to_number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = normalize_text(value).replace("%", "")
    try:
        return float(text)
    except ValueError:
        return 0.0


def to_int(value: Any) -> int:
    return int(round(to_number(value)))


def canonicalize_work_center_id(value: Any) -> str:
    text = normalize_text(value).upper()
    match = re.match(r"^WC-?(\d{1,2})$", text)
    if match:
        return f"WC-{int(match.group(1)):02d}"
    return text


def relative_display_path(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(BASE_DIR))
    except ValueError:
        return str(path.resolve())


def extract_series(row: tuple[Any, ...] | list[Any], start_index: int, length: int) -> list[float]:
    return [to_number(row[start_index + offset] if len(row) > start_index + offset else None) for offset in range(length)]


def split_item_label(label: Any) -> tuple[str, str]:
    text = normalize_text(label)
    if not text:
        return "", ""

    match = re.match(r"^([A-Z]{1,4}-\d{2,3}[A-Za-z]?)\s*(?:[|:\-\u2013\u2014])?\s*(.*)$", text)
    if match:
        code = match.group(1)
        name = match.group(2).strip(" -|:\u2013\u2014")
        return code, name or code

    return "", text


def split_work_center_label(label: Any) -> tuple[str, str]:
    text = normalize_text(label)
    if not text:
        return "", ""

    match = re.match(r"^(WC-?\d{1,2})\s*(?:[|:\-\u2013\u2014])?\s*(.*)$", text, flags=re.IGNORECASE)
    if match:
        work_center_id = canonicalize_work_center_id(match.group(1))
        name = match.group(2).strip(" -|:\u2013\u2014")
        return work_center_id, name or work_center_id

    return text, text


def get_visible_sheet(workbook, name: str | tuple[str, ...] | list[str]):
    candidate_names = [name] if isinstance(name, str) else list(name)
    selected_name = next((candidate for candidate in candidate_names if candidate in workbook.sheetnames), None)
    if selected_name is None:
        names_label = " / ".join(candidate_names)
        raise KeyError(f'Sheet "{names_label}" tidak ditemukan di workbook.')

    sheet = workbook[selected_name]
    if getattr(sheet, "sheet_state", "visible") != "visible":
        raise ValueError(f'Sheet "{selected_name}" sedang hidden dan tidak boleh dipakai sebagai sumber data.')

    return sheet


def get_optional_visible_sheet(workbook, name: str | tuple[str, ...] | list[str]):
    try:
        return get_visible_sheet(workbook, name)
    except KeyError:
        return None


def list_available_workbooks() -> list[dict[str, Any]]:
    files: list[Path] = []

    if DATA_DIR.exists():
        for path in sorted(DATA_DIR.glob("*.xlsx"), key=lambda item: item.stat().st_mtime_ns, reverse=True):
            if path.name.startswith("~$"):
                continue
            files.append(path.resolve())

    return [
        {
            "name": path.name,
            "path": relative_display_path(path),
            "size_bytes": path.stat().st_size,
        }
        for path in files
    ]


def resolve_workbook_path(workbook: str | None = None) -> Path:
    if workbook:
        candidate = Path(workbook)
        if not candidate.is_absolute():
            candidate = (DATA_DIR / workbook).resolve() if (DATA_DIR / workbook).exists() else (BASE_DIR / workbook).resolve()
        if not candidate.exists():
            raise FileNotFoundError(f'File workbook "{workbook}" tidak ditemukan.')
        return candidate

    available = list_available_workbooks()
    if not available:
        raise FileNotFoundError(
            'Tidak ada file Excel input. Simpan file ".xlsx" ke folder "Data TA".'
        )

    return (BASE_DIR / available[0]["path"]).resolve()


def parse_product_families(sheet) -> dict[str, dict[str, str]]:
    family_name_map: dict[str, str] = {}
    product_map: dict[str, dict[str, str]] = {}

    for row in sheet.iter_rows(values_only=True):
        first = normalize_text(row[0] if len(row) > 0 else None)
        second = normalize_text(row[1] if len(row) > 1 else None)
        third = normalize_text(row[2] if len(row) > 2 else None)

        if first.startswith("PRD-"):
            family_name_map[first] = second
        elif first.startswith("ROTI-"):
            try:
                code = f'P-{int(first.split("-")[1]):02d}'
            except (IndexError, ValueError):
                continue
            product_map[code] = {
                "family_id": third,
                "family_name": family_name_map.get(third, third),
            }

    return product_map


def parse_products(sheet, product_families: dict[str, dict[str, str]]) -> tuple[list[int], list[dict[str, Any]]]:
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 4:
        raise ValueError("Sheet Master Production Schedule tidak memiliki struktur yang valid.")

    week_numbers = [to_int(value) for value in rows[2][2:] if value is not None]
    products: list[dict[str, Any]] = []

    for row in rows[3:]:
        name = normalize_text(row[1] if len(row) > 1 else None)
        values = [to_int(value) for value in row[2 : 2 + len(week_numbers)]]
        if not name or not any(value != 0 for value in values):
            continue

        product_id = f'P-{len(products) + 1:02d}'
        family_info = product_families.get(product_id, {})
        products.append(
            {
                "id": product_id,
                "name": name.title(),
                "family_id": family_info.get("family_id", ""),
                "family_name": family_info.get("family_name", ""),
                "price": 0,
                "color": PRODUCT_COLORS[len(products) % len(PRODUCT_COLORS)],
                "weekly_plan": values,
            }
        )

    return week_numbers, products


def build_unique_product_families(products: list[dict[str, Any]]) -> list[dict[str, Any]]:
    families: list[dict[str, Any]] = []
    seen: set[str] = set()

    for product in products:
        family_id = product.get("family_id") or ""
        if not family_id or family_id in seen:
            continue
        seen.add(family_id)
        families.append(
            {
                "id": family_id,
                "name": product.get("family_name") or family_id,
                "color": product.get("color") or PRODUCT_COLORS[len(families) % len(PRODUCT_COLORS)],
            }
        )

    return families


def parse_aggregate_demand(sheet, products: list[dict[str, Any]], periods_count: int) -> dict[str, Any]:
    rows = list(sheet.iter_rows(values_only=True))
    header_index = -1
    product_columns: list[tuple[int, str]] = []

    for index, row in enumerate(rows):
        if normalize_key(row[1] if len(row) > 1 else None) != "periode":
            continue
        for column_index in range(2, len(row)):
            label = normalize_text(row[column_index])
            if label:
                product_columns.append((column_index, label))
        if product_columns:
            header_index = index
            break

    if header_index == -1 or not product_columns:
        return {"rows": [], "periods": [], "source_sheet": sheet.title}

    families = build_unique_product_families(products)
    forecast_rows: list[dict[str, Any]] = []

    for offset, (column_index, label) in enumerate(product_columns):
        family = families[offset] if offset < len(families) else {}
        product_id = family.get("id") or f"PRD-{offset + 1:02d}"
        values: list[float] = []

        for row in rows[header_index + 1 :]:
            period_no = to_int(row[1] if len(row) > 1 else None)
            if period_no <= 0:
                continue
            values.append(to_number(row[column_index] if len(row) > column_index else None))
            if len(values) >= periods_count:
                break

        if len(values) < periods_count:
            values.extend([0.0] * (periods_count - len(values)))

        forecast_rows.append(
            {
                "id": product_id,
                "name": family.get("name") or label,
                "label": label,
                "color": family.get("color") or PRODUCT_COLORS[offset % len(PRODUCT_COLORS)],
                "values": values[:periods_count],
                "total": round(sum(values[:periods_count]), 2),
            }
        )

    periods = []
    for index in range(periods_count):
        period_row = {
            "week": index + 1,
            "week_number": index + 1,
            "period": f"W{index + 1}",
            "total": 0.0,
        }
        for row in forecast_rows:
            value = row["values"][index] if index < len(row["values"]) else 0.0
            period_row[row["id"]] = value
            period_row["total"] += value
        period_row["total"] = round(period_row["total"], 2)
        periods.append(period_row)

    return {
        "rows": forecast_rows,
        "periods": periods,
        "source_sheet": sheet.title,
        "total": round(sum(row["total"] for row in forecast_rows), 2),
    }


def parse_ingredient_master(sheet) -> dict[str, dict[str, Any]]:
    ingredient_map: dict[str, dict[str, Any]] = {}

    for row in sheet.iter_rows(min_row=5, values_only=True):
        code = normalize_text(row[2] if len(row) > 2 else None)
        name = normalize_text(row[3] if len(row) > 3 else None)
        if not name:
            continue

        payload = {
            "code": code,
            "name": name,
            "category": normalize_text(row[4] if len(row) > 4 else None),
            "producer": normalize_text(row[5] if len(row) > 5 else None),
            "country": normalize_text(row[6] if len(row) > 6 else None),
            "supplier": normalize_text(row[7] if len(row) > 7 else None),
        }

        keys = {
            normalize_key(name),
            normalize_key(name.split("(")[0]),
        }

        for key in keys:
            if key and key not in ingredient_map:
                ingredient_map[key] = payload

    return ingredient_map


def parse_bom(
    sheet,
    ingredient_map: dict[str, dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, list[dict[str, Any]]]]:
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 15:
        raise ValueError("Sheet Matriks Bahan Baku tidak memiliki struktur yang valid.")

    header = rows[13]
    product_columns: list[tuple[int, str]] = []
    for index in range(6, len(header)):
        label = normalize_text(header[index])
        match = re.search(r"(P-\d{2})", label)
        if match:
            product_columns.append((index, match.group(1)))

    components: list[dict[str, Any]] = []
    components_by_id: dict[str, dict[str, Any]] = {}
    bom_entries: list[dict[str, Any]] = []
    bom_by_product: dict[str, list[dict[str, Any]]] = {code: [] for _, code in product_columns}

    for row in rows[14:]:
        component_name = normalize_text(row[4] if len(row) > 4 else None)
        unit = normalize_text(row[5] if len(row) > 5 else None)
        if not component_name or not unit:
            continue

        meta = ingredient_map.get(normalize_key(component_name), {})
        component_id = f"CMP-{slugify(component_name)}"
        if component_id not in components_by_id:
            component_payload = {
                "id": component_id,
                "name": component_name,
                "unit": unit,
                "code": meta.get("code", ""),
                "category": meta.get("category", ""),
                "supplier": meta.get("supplier", ""),
                "producer": meta.get("producer", ""),
            }
            components_by_id[component_id] = component_payload
            components.append(component_payload)

        for column_index, product_id in product_columns:
            quantity = to_number(row[column_index] if len(row) > column_index else None)
            if quantity <= 0:
                continue

            entry = {
                "component_id": component_id,
                "component_name": component_name,
                "unit": unit,
                "product_id": product_id,
                "quantity_per_unit": quantity,
            }
            bom_entries.append(entry)
            bom_by_product.setdefault(product_id, []).append(
                {
                    "id": component_id,
                    "name": component_name,
                    "unit": unit,
                    "quantity_per_unit": quantity,
                    "code": components_by_id[component_id]["code"],
                }
            )

    for product_components in bom_by_product.values():
        product_components.sort(key=lambda item: item["name"])

    return components, bom_entries, bom_by_product


def parse_routing(sheet) -> tuple[list[dict[str, Any]], dict[str, list[dict[str, Any]]], dict[str, list[dict[str, Any]]]]:
    routing_entries: list[dict[str, Any]] = []
    routing_by_product: dict[str, list[dict[str, Any]]] = {}
    routing_by_item: dict[str, list[dict[str, Any]]] = {}
    current_item_label = ""

    for row in sheet.iter_rows(min_row=5, values_only=True):
        item_label = normalize_text(row[0] if len(row) > 0 else None)
        level = normalize_text(row[1] if len(row) > 1 else None)
        work_center_id = normalize_text(row[2] if len(row) > 2 else None)

        if normalize_key(item_label) == "workcentermasterfile":
            break
        if item_label.startswith("*"):
            continue

        if item_label:
            current_item_label = item_label

        if not current_item_label or not work_center_id.startswith("WC-"):
            continue

        item_code, item_name = split_item_label(current_item_label)
        item_key = item_code or current_item_label
        entry = {
            "item_key": item_key,
            "item_code": item_code,
            "item_name": item_name,
            "item_label": current_item_label,
            "level": level,
            "work_center_id": work_center_id,
            "setup_minutes": to_number(row[3] if len(row) > 3 else None),
            "run_minutes": to_number(row[4] if len(row) > 4 else None),
            "description": normalize_text(row[5] if len(row) > 5 else None),
        }
        routing_entries.append(entry)
        routing_by_item.setdefault(item_key, []).append(entry)
        if item_code.startswith("P-"):
            routing_by_product.setdefault(item_code, []).append(entry)

    return routing_entries, routing_by_product, routing_by_item


def parse_work_centers(sheet) -> list[dict[str, Any]]:
    work_centers: list[dict[str, Any]] = []

    for row in sheet.iter_rows(values_only=True):
        label = normalize_text(row[0] if len(row) > 0 else None)
        work_center_id, name = split_work_center_label(label)
        if not work_center_id.startswith("WC-"):
            continue

        capacity_hours = to_number(row[5] if len(row) > 5 else None)
        work_centers.append(
            {
                "id": work_center_id,
                "name": name,
                "units": to_int(row[1] if len(row) > 1 else None),
                "utilization_label": normalize_text(row[2] if len(row) > 2 else None),
                "queue_days": to_number(row[3] if len(row) > 3 else None),
                "hours_per_day": to_number(row[4] if len(row) > 4 else None),
                "capacity_hours": capacity_hours,
                "capacity_minutes": round(capacity_hours * 60, 2),
                "notes": normalize_text(row[6] if len(row) > 6 else None),
            }
        )

    return work_centers


def parse_mrp_sheet(
    sheet,
    periods_count: int,
    product_lookup_by_name: dict[str, dict[str, Any]],
    ingredient_lookup_by_name: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    rows = list(sheet.iter_rows(values_only=True))
    items: list[dict[str, Any]] = []
    items_by_category = {
        "final_product": [],
        "intermediate": [],
        "raw_material": [],
    }

    for row_index, row in enumerate(rows):
        if normalize_text(row[4] if len(row) > 4 else None) != "Item":
            continue

        item_label = normalize_text(row[5] if len(row) > 5 else None)
        if not item_label:
            continue

        code, parsed_name = split_item_label(item_label)
        name_key = normalize_key(parsed_name)
        product_match = product_lookup_by_name.get(name_key)
        lot_size_value = rows[row_index + 1][5] if row_index + 1 < len(rows) and len(rows[row_index + 1]) > 5 else None
        lead_time_value = rows[row_index + 1][7] if row_index + 1 < len(rows) and len(rows[row_index + 1]) > 7 else None
        ingredient_match = ingredient_lookup_by_name.get(name_key)

        if product_match:
            category = "final_product"
        elif (code and code.startswith("B-")) or ingredient_match:
            category = "raw_material"
        else:
            category = "intermediate"

        item_code = code or (product_match["id"] if product_match else "")
        item_id = f"mrp-{slugify(item_code or item_label)}"
        gross_requirement = extract_series(rows[row_index + 2], 8, periods_count)
        scheduled_receipt = extract_series(rows[row_index + 3], 8, periods_count)
        projected_on_hand = extract_series(rows[row_index + 4], 8, periods_count)
        net_requirement = extract_series(rows[row_index + 5], 8, periods_count)
        planned_order_receipt = extract_series(rows[row_index + 6], 8, periods_count)
        planned_order_release = extract_series(rows[row_index + 7], 8, periods_count)

        item_payload = {
            "id": item_id,
            "code": item_code,
            "name": parsed_name,
            "label": item_label,
            "category": category,
            "level": normalize_text(row[6] if len(row) > 6 else None),
            "lot_size": normalize_text(lot_size_value) if isinstance(lot_size_value, str) else to_number(lot_size_value),
            "lead_time": to_int(lead_time_value),
            "product_id": product_match["id"] if product_match else "",
            "color": product_match["color"] if product_match else "",
            "gross_requirement_values": gross_requirement,
            "scheduled_receipt_values": scheduled_receipt,
            "projected_on_hand_values": projected_on_hand,
            "net_requirement_values": net_requirement,
            "planned_order_receipt_values": planned_order_receipt,
            "planned_order_release_values": planned_order_release,
            "total_gross_requirement": round(sum(gross_requirement), 2),
            "total_scheduled_receipt": round(sum(scheduled_receipt), 2),
            "total_projected_on_hand": round(sum(projected_on_hand), 2),
            "total_net_requirement": round(sum(net_requirement), 2),
            "total_planned_order_receipt": round(sum(planned_order_receipt), 2),
            "total_planned_order_release": round(sum(planned_order_release), 2),
        }
        items.append(item_payload)
        items_by_category[category].append(item_payload)

    item_lookup = {item["id"]: item for item in items}
    return {
        "items": items,
        "item_lookup": item_lookup,
        "by_category": items_by_category,
    }


def parse_crp_detail_rows(rows: list[tuple[Any, ...]], start_index: int, end_index: int, periods_count: int, row_type: str) -> list[dict[str, Any]]:
    detail_rows: list[dict[str, Any]] = []
    current_work_center_id = ""
    current_work_center_name = ""

    for row in rows[start_index:end_index]:
        work_center_label = normalize_text(row[1] if len(row) > 1 else None)
        item_label = normalize_text(row[2] if len(row) > 2 else None)
        parsed_work_center_id, parsed_work_center_name = split_work_center_label(work_center_label)
        if parsed_work_center_id.startswith("WC-"):
            current_work_center_id = parsed_work_center_id
            current_work_center_name = parsed_work_center_name
        if not current_work_center_id or not item_label:
            continue

        item_code, item_name = split_item_label(item_label)
        values = extract_series(row, 3, periods_count)
        detail_rows.append(
            {
                "type": row_type,
                "work_center_id": current_work_center_id,
                "work_center_name": current_work_center_name,
                "item_code": item_code,
                "item_name": item_name,
                "item_label": item_label,
                "values": values,
                "total_minutes": round(sum(values), 2),
            }
        )

    return detail_rows


def parse_crp_sheet(sheet, periods_count: int) -> dict[str, Any]:
    rows = list(sheet.iter_rows(values_only=True))

    run_header_index = -1
    setup_marker_index = -1
    total_marker_index = -1
    total_header_index = -1
    capacity_marker_index = -1
    capacity_header_index = -1

    for index, row in enumerate(rows):
        second = normalize_text(row[1] if len(row) > 1 else None)
        third = normalize_text(row[2] if len(row) > 2 else None)
        if run_header_index == -1 and second == "WC" and third == "Item":
            run_header_index = index
        elif second == "Setup time matrix":
            setup_marker_index = index
        elif second == "Kebutuhan kapasitas PORel":
            total_marker_index = index
        elif total_marker_index != -1 and total_header_index == -1 and second == "WC" and third == "Keterangan":
            total_header_index = index
        elif second == "Capacity Requirement Planning (CRP)":
            capacity_marker_index = index
        elif capacity_marker_index != -1 and capacity_header_index == -1 and second == "WC" and third == "Periode":
            capacity_header_index = index

    if run_header_index == -1 or setup_marker_index == -1 or total_header_index == -1:
        raise ValueError("Sheet CRP tidak memiliki struktur visible yang valid.")

    run_rows = parse_crp_detail_rows(rows, run_header_index + 1, setup_marker_index, periods_count, "run")
    setup_rows = parse_crp_detail_rows(rows, setup_marker_index + 1, total_marker_index, periods_count, "setup")
    work_center_id_by_name = {
        normalize_key(row["work_center_name"]): row["work_center_id"]
        for row in [*run_rows, *setup_rows]
        if row.get("work_center_name") and row.get("work_center_id")
    }

    totals_by_wc: dict[str, dict[str, Any]] = {}
    available_time_by_wc: dict[str, list[float]] = {}
    total_header = rows[total_header_index]
    available_time_index = next(
        (
            column_index
            for column_index, value in enumerate(total_header)
            if normalize_key(value) == "availabletimeweek"
        ),
        -1,
    )
    for row in rows[total_header_index + 1 :]:
        work_center_label = normalize_text(row[1] if len(row) > 1 else None)
        description = normalize_text(row[2] if len(row) > 2 else None)
        if not work_center_label.startswith("WC-"):
            if totals_by_wc:
                break
            continue

        work_center_id, work_center_name = split_work_center_label(work_center_label)
        work_center_id = work_center_id_by_name.get(normalize_key(work_center_name), work_center_id)
        values = extract_series(row, 3, periods_count)
        totals_by_wc[work_center_id] = {
            "id": work_center_id,
            "name": work_center_name,
            "description": description,
            "values": values,
            "total_minutes": round(sum(values), 2),
        }
        weekly_capacity = to_number(row[available_time_index] if available_time_index != -1 and len(row) > available_time_index else None)
        if weekly_capacity:
            available_time_by_wc[work_center_id] = [weekly_capacity] * periods_count

    if capacity_header_index != -1:
        capacity_header = rows[capacity_header_index]
        capacity_week_index = next(
            (
                column_index
                for column_index, value in enumerate(capacity_header)
                if normalize_key(value) == "capacityweek"
            ),
            -1,
        )
        for row in rows[capacity_header_index + 2 :]:
            work_center_id = canonicalize_work_center_id(row[1] if len(row) > 1 else None)
            if not work_center_id.startswith("WC-"):
                if available_time_by_wc:
                    break
                continue
            values = extract_series(row, 2, periods_count)
            weekly_capacity = to_number(row[capacity_week_index] if capacity_week_index != -1 and len(row) > capacity_week_index else None)
            if any(values):
                available_time_by_wc[work_center_id] = values
            elif weekly_capacity:
                available_time_by_wc[work_center_id] = [weekly_capacity] * periods_count

    return {
        "run_rows": run_rows,
        "setup_rows": setup_rows,
        "totals_by_wc": totals_by_wc,
        "available_time_by_wc": available_time_by_wc,
    }


def parse_bol_rccp_sheet(sheet, periods_count: int) -> dict[str, Any]:
    rows = list(sheet.iter_rows(values_only=True))
    bill_of_labour_by_wc: dict[str, dict[str, Any]] = {}
    loads_by_wc: dict[str, list[float]] = {}
    available_time_by_wc: dict[str, list[float]] = {}
    lot_size_by_item: dict[str, float] = {}
    lot_size_by_name: dict[str, float] = {}
    rccp_week_numbers: list[int] = []
    work_center_order: list[str] = []

    bol_detail_header_index = -1
    bill_header_index = -1
    rccp_marker_index = -1
    rccp_header_index = -1

    for index, row in enumerate(rows):
        second = normalize_text(row[1] if len(row) > 1 else None)
        third = normalize_text(row[2] if len(row) > 2 else None)
        fourth = normalize_text(row[3] if len(row) > 3 else None)
        ninth = normalize_text(row[8] if len(row) > 8 else None)
        if bol_detail_header_index == -1 and second == "Item" and third == "Level" and fourth == "Work Center" and ninth == "Lot Size":
            bol_detail_header_index = index
        if bill_header_index == -1 and second == "Work Center" and third == "Processing Time":
            bill_header_index = index
        if second == "RCCP":
            rccp_marker_index = index
        elif rccp_marker_index != -1 and rccp_header_index == -1 and second == "Work Center" and third == "Week":
            rccp_header_index = index
            break

    if bill_header_index == -1 or rccp_header_index == -1:
        raise ValueError('Sheet "BOL + RCCP" tidak memiliki struktur RCCP yang valid.')

    if bol_detail_header_index != -1:
        for row in rows[bol_detail_header_index + 1 : bill_header_index]:
            item_label = normalize_text(row[1] if len(row) > 1 else None)
            lot_size = to_number(row[8] if len(row) > 8 else None)
            if not item_label or lot_size <= 0:
                continue
            item_code, item_name = split_item_label(item_label)
            if item_code:
                lot_size_by_item[item_code] = lot_size
            if item_name:
                lot_size_by_name[normalize_key(item_name)] = lot_size
            lot_size_by_name[normalize_key(item_label)] = lot_size

    for row in rows[bill_header_index + 1 :]:
        work_center_id = canonicalize_work_center_id(row[1] if len(row) > 1 else None)
        if not work_center_id.startswith("WC-"):
            if bill_of_labour_by_wc:
                break
            continue
        bill_of_labour_by_wc[work_center_id] = {
            "id": work_center_id,
            "processing_time": round(to_number(row[2] if len(row) > 2 else None), 2),
        }

    rccp_week_numbers = [
        to_int(value)
        for value in rows[rccp_header_index + 1][2 : 2 + periods_count]
    ]

    load_end_index = rccp_header_index + 2
    for row in rows[rccp_header_index + 2 :]:
        work_center_id = canonicalize_work_center_id(row[1] if len(row) > 1 else None)
        if not work_center_id.startswith("WC-"):
            break
        work_center_order.append(work_center_id)
        loads_by_wc[work_center_id] = extract_series(row, 2, periods_count)
        load_end_index += 1

    available_start_index = load_end_index
    blank_seen = False
    for index in range(load_end_index, len(rows)):
        work_center_id = canonicalize_work_center_id(rows[index][1] if len(rows[index]) > 1 else None)
        if work_center_id.startswith("WC-"):
            if blank_seen:
                available_start_index = index
                break
        elif any(normalize_text(value) for value in rows[index]):
            continue
        else:
            blank_seen = True

    for row in rows[available_start_index:]:
        work_center_id = canonicalize_work_center_id(row[1] if len(row) > 1 else None)
        if not work_center_id.startswith("WC-"):
            if available_time_by_wc:
                break
            continue
        available_time_by_wc[work_center_id] = extract_series(row, 2, periods_count)

    return {
        "week_numbers": rccp_week_numbers,
        "work_center_order": work_center_order,
        "bill_of_labour_by_wc": bill_of_labour_by_wc,
        "loads_by_wc": loads_by_wc,
        "available_time_by_wc": available_time_by_wc,
        "lot_size_by_item": lot_size_by_item,
        "lot_size_by_name": lot_size_by_name,
    }


def apply_bol_lot_sizes_to_mrp(mrp: dict[str, Any], bol_rccp: dict[str, Any]) -> None:
    lot_size_by_item = bol_rccp.get("lot_size_by_item", {})
    lot_size_by_name = bol_rccp.get("lot_size_by_name", {})

    for item in mrp.get("items", []):
        lot_size = lot_size_by_item.get(item.get("code") or "")
        if lot_size is None:
            lot_size = lot_size_by_name.get(normalize_key(item.get("name")))
        if lot_size is None:
            lot_size = lot_size_by_name.get(normalize_key(item.get("label")))
        if lot_size is not None:
            item["lot_size"] = lot_size


def parse_routes(route_sheet, route_store_sheet) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    routes: list[dict[str, Any]] = []
    route_lookup: dict[str, dict[str, Any]] = {}

    for row in route_sheet.iter_rows(min_row=2, values_only=True):
        route_id = normalize_text(row[0] if len(row) > 0 else None)
        if not route_id.startswith("RTE"):
            continue

        route = {
            "id": route_id,
            "start_location_id": normalize_text(row[1] if len(row) > 1 else None),
            "end_location_id": normalize_text(row[2] if len(row) > 2 else None),
            "name": normalize_text(row[3] if len(row) > 3 else None),
            "day": normalize_text(row[4] if len(row) > 4 else None),
            "color": DAY_COLORS.get(normalize_text(row[4] if len(row) > 4 else None), "#64748b"),
            "stores": [],
            "store_count": 0,
        }
        routes.append(route)
        route_lookup[route_id] = route

    stores: list[dict[str, Any]] = []
    for row in route_store_sheet.iter_rows(min_row=2, values_only=True):
        store_id = normalize_text(row[1] if len(row) > 1 else None)
        if not store_id:
            continue

        route_ids = [
            route_id.strip()
            for route_id in normalize_text(row[4] if len(row) > 4 else None).split(",")
            if route_id.strip()
        ]
        store = {
            "id": store_id,
            "name": normalize_text(row[2] if len(row) > 2 else None),
            "location": normalize_text(row[3] if len(row) > 3 else None),
            "route_ids": route_ids,
        }
        stores.append(store)

        for route_id in route_ids:
            route = route_lookup.get(route_id)
            if route is None:
                continue
            route["stores"].append(
                {
                    "id": store["id"],
                    "name": store["name"],
                    "location": store["location"],
                }
            )

    for route in routes:
        route["stores"].sort(key=lambda item: item["name"])
        route["store_count"] = len(route["stores"])

    return routes, stores


def parse_forward_schedule_sheet(sheet, periods_count: int) -> dict[str, Any]:
    rows = list(sheet.iter_rows(values_only=True))
    header_index = -1

    for index, row in enumerate(rows):
        first = normalize_key(row[0] if len(row) > 0 else None)
        third = normalize_key(row[2] if len(row) > 2 else None)
        if first.startswith("produk") and third.startswith("periode"):
            header_index = index
            break

    if header_index == -1:
        return {"rows": [], "source_sheet": sheet.title}

    schedule_rows: list[dict[str, Any]] = []
    operation_index_by_product_period: dict[tuple[str, int], int] = {}

    for row in rows[header_index + 1 :]:
        product_label = normalize_text(row[0] if len(row) > 0 else None)
        match = re.match(r"^(P-\d{2})\s*(.*)$", product_label)
        if not match:
            continue

        period_no = to_int(row[2] if len(row) > 2 else None)
        if period_no <= 0 or period_no > periods_count:
            continue

        quantity = to_number(row[3] if len(row) > 3 else None)
        if quantity <= 0:
            continue

        product_id = match.group(1)
        product_name = normalize_text(match.group(2)) or product_id
        group_key = (product_id, period_no)
        operation_no = operation_index_by_product_period.get(group_key, 0) + 1
        operation_index_by_product_period[group_key] = operation_no

        schedule_rows.append(
            {
                "product_id": product_id,
                "product_name": product_name,
                "period": period_no,
                "quantity": quantity,
                "lot_size": to_number(row[1] if len(row) > 1 else None),
                "lot_count": to_number(row[4] if len(row) > 4 else None),
                "level": normalize_text(row[5] if len(row) > 5 else None),
                "work_center_id": canonicalize_work_center_id(row[6] if len(row) > 6 else None),
                "work_center_name": normalize_text(row[7] if len(row) > 7 else None),
                "operation_no": operation_no,
                "operation_name": normalize_text(row[8] if len(row) > 8 else None),
                "setup_minutes": to_number(row[9] if len(row) > 9 else None),
                "run_minutes": to_number(row[10] if len(row) > 10 else None),
                "duration_minutes": to_number(row[11] if len(row) > 11 else None),
                "duration_hours": to_number(row[12] if len(row) > 12 else None),
                "start_hour": to_number(row[13] if len(row) > 13 else None),
                "finish_hour": to_number(row[14] if len(row) > 14 else None),
                "status_label": normalize_text(row[15] if len(row) > 15 else None),
            }
        )

    return {
        "rows": schedule_rows,
        "source_sheet": sheet.title,
        "available_periods": len({row["period"] for row in schedule_rows}),
    }


@lru_cache(maxsize=4)
def _load_dataset(path_str: str, mtime_ns: int, size_bytes: int) -> dict[str, Any]:
    workbook_path = Path(path_str)
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)

    visible_sheets = [worksheet.title for worksheet in workbook.worksheets if worksheet.sheet_state == "visible"]

    data_produk_sheet = get_visible_sheet(workbook, "Data Produk")
    mps_sheet = get_visible_sheet(workbook, "Master Production Schedule")
    ingredient_sheet = get_visible_sheet(workbook, "Daftar Nama Bahan")
    bom_sheet = get_visible_sheet(workbook, "Matriks Bahan Baku")
    routing_sheet = get_visible_sheet(workbook, "1. Routing & WC Files")
    route_sheet = get_visible_sheet(workbook, "Daftar Rute")
    route_store_sheet = get_visible_sheet(workbook, "Rute-Toko")
    mrp_sheet = get_visible_sheet(workbook, "MRP")
    crp_sheet = get_visible_sheet(workbook, "CRP")
    bol_rccp_sheet = get_visible_sheet(workbook, ("BOL + RCCP", "RCCP + BOL"))
    aggregate_demand_sheet = get_optional_visible_sheet(workbook, ("Agregate Demand", "Aggregate Demand"))
    forward_schedule_sheet = get_optional_visible_sheet(workbook, ("Forward Scheduling (2)", "Forward Scheduling"))

    product_families = parse_product_families(data_produk_sheet)
    week_numbers, products = parse_products(mps_sheet, product_families)
    forecast = (
        parse_aggregate_demand(aggregate_demand_sheet, products, len(week_numbers))
        if aggregate_demand_sheet is not None
        else {"rows": [], "periods": [], "source_sheet": ""}
    )
    ingredient_map = parse_ingredient_master(ingredient_sheet)
    components, bom_entries, bom_by_product = parse_bom(bom_sheet, ingredient_map)
    routing_entries, routing_by_product, routing_by_item = parse_routing(routing_sheet)
    work_centers = parse_work_centers(routing_sheet)
    routes, stores = parse_routes(route_sheet, route_store_sheet)

    product_lookup = {product["id"]: product for product in products}
    product_lookup_by_name = {normalize_key(product["name"]): product for product in products}
    work_center_lookup = {work_center["id"]: work_center for work_center in work_centers}

    mrp = parse_mrp_sheet(mrp_sheet, len(week_numbers), product_lookup_by_name, ingredient_map)
    crp = parse_crp_sheet(crp_sheet, len(week_numbers))
    bol_rccp = parse_bol_rccp_sheet(bol_rccp_sheet, len(week_numbers))
    production_schedule = (
        parse_forward_schedule_sheet(forward_schedule_sheet, len(week_numbers))
        if forward_schedule_sheet is not None
        else {"rows": [], "source_sheet": ""}
    )
    apply_bol_lot_sizes_to_mrp(mrp, bol_rccp)

    return {
        "source_name": workbook_path.name,
        "source_path": relative_display_path(workbook_path),
        "visible_sheets": visible_sheets,
        "available_periods": len(week_numbers),
        "week_numbers": week_numbers,
        "products": products,
        "forecast": forecast,
        "product_lookup": product_lookup,
        "components": components,
        "bom_entries": bom_entries,
        "bom_by_product": bom_by_product,
        "routing_entries": routing_entries,
        "routing_by_product": routing_by_product,
        "routing_by_item": routing_by_item,
        "work_centers": work_centers,
        "work_center_lookup": work_center_lookup,
        "mrp": mrp,
        "bol_rccp": bol_rccp,
        "crp": crp,
        "production_schedule": production_schedule,
        "routes": routes,
        "stores": stores,
    }


def get_dataset(workbook: str | None = None) -> dict[str, Any]:
    path = resolve_workbook_path(workbook)
    stat = path.stat()
    return _load_dataset(str(path.resolve()), stat.st_mtime_ns, stat.st_size)
